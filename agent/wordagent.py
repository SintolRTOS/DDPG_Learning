# -*- coding: utf-8 -*-
"""
Created on Tue May 21 16:04:40 2019

@author: wangjingyi
"""

import sys 
sys.path.append("..") 

from collections import OrderedDict
import numpy as np
np.set_printoptions(suppress=True)
from openpyxl import load_workbook
import logger
from agent.box import Box
from agent.dict import Dict
from agent.rank import KeyWordRank
from gym.utils import seeding
import agent.reward


#高流量
COLLAGE_HIGHFLOW = 1
#高转化
COLLAGE_HIGHCONVERSION = 2
#高ROI
COLLAGE_HIGHROI = 3
#高转化+高ROI
COLLAGE_HIGHROI_HIGHCONVERSION = 4

POPULARITY_BOUND = 1000000
HIGH_CONVERTION_PARAMS_SIZE = 5
HIGH_FLOW_PARAMS_SIZE = 3
HIGH_ROI_PARAMS_SIZE = 3
HIGH_ROI_CONVERTION_PARAMS_SIZE = 8
ALL_WORDS_MAX = 20
rank = KeyWordRank()


class WordAgent(object):
    """docstring for ClassName"""
    def __init__(self,
                 filepath,
                 mode,
                 reward_type = COLLAGE_HIGHCONVERSION):
        super(WordAgent,self).__init__()
        self.filepath = filepath
        self.mode = mode
        self.reward_type = int(reward_type);

        if self.reward_type == COLLAGE_HIGHCONVERSION:
            self.parameter_size = HIGH_CONVERTION_PARAMS_SIZE
            
        elif self.reward_type == COLLAGE_HIGHFLOW:
            self.parameter_size = HIGH_FLOW_PARAMS_SIZE
            
        elif self.reward_type == COLLAGE_HIGHROI:
            self.parameter_size = HIGH_ROI_PARAMS_SIZE
            
        elif self.reward_type == COLLAGE_HIGHROI_HIGHCONVERSION:
            self.parameter_size = HIGH_ROI_CONVERTION_PARAMS_SIZE
        
        self.result = []
        self.all_words = ALL_WORDS_MAX
        self.result_keywords = []
        self.select_index = 0;
        self.select_count = 0;
        self.select_process = 0.
        self.select_total = 0
        self.select_total_reward = 0.
        self.max_action = 1.
        self.reward_range = (-float('inf'),float('inf'))
        self.metadata = {'render.modes':[]}
        self.spec = None
        self.openExcel()
        self.reset()
        
    def openExcel(self):
        logger.debug('start openpyxl openExcel!')
        self.wb = load_workbook(self.filepath)
        logger.debug('form sheetname: ' + self.wb.sheetnames[0])
        self.keytitle = self.wb.sheetnames[0];
        sheet = self.wb.get_sheet_by_name(str(self.keytitle))
        self.max_row = sheet.max_row
        self.max_column = sheet.max_column
        logger.debug("form max_row: " + str(self.max_row))
        logger.debug("form max_column: " + str(self.max_column))
        self.data = []
        #wish data base
        for row in sheet.iter_rows(min_col=1, min_row=1, max_row=self.max_row, max_col=self.max_column):
            ROW = []
            if row[1].value == 0:
                continue;
            if row[2].value == 0:
                continue;
            if row[9].value == 0 and row[10].value == 0 and row[11].value == 0:
                continue;
            for cell in row:
                ROW.append(cell.value)
            self.data.append(ROW)
        logger.debug("totole data row: " + str(len(self.data)))
        logger.debug("all form data: ")
        logger.debug(self.data)
    
    def print_result(self):
        if len(self.result) == 0:
            return
        keyword = self.get_result_keywords()
        rank.checkrank(self.select_total_reward,keyword)
        logger.info('----------------print_result------------------')
        renderdic = {}
        renderdic['select_total_reward：'] = self.select_total_reward
        renderdic['result_idlist：'] = self.result
        renderdic['result_keywords：'] = keyword
        logger.info(str(renderdic))
        logger.info('----------------print_result------------------')
        
        logger.info('print_rank_start')
        logger.info(rank.rank_value)
        logger.info('current rank value: ')
        index = len(rank.rank_list) -1
        for n in range(len(rank.rank_list)):
            logger.info(rank.rank_list[index])
            index = index - 1
        logger.info('current rank words: ')
        logger.info('print_rank_end')
        
    def reset(self):
        self.print_result()
        
        observation = None

        if self.reward_type == COLLAGE_HIGHCONVERSION:
            logger.info('init COLLAGE_HIGHCONVERSION')
            observation = self.init_observation()
        elif self.reward_type == COLLAGE_HIGHFLOW:
            logger.info('init COLLAGE_HIGHFLOW')
            observation = self.init_observation_highflow()
        elif self.reward_type == COLLAGE_HIGHROI:
            logger.info('init COLLAGE_HIGHROI')
            observation = self.init_observation_highroi()
        elif self.reward_type == COLLAGE_HIGHROI_HIGHCONVERSION:
            logger.info('init COLLAGE_HIGHROI_HIGHCONVERSION')
            observation = self.init_observation_highroi_hightconveration()
        
        if observation is False:
            logger.error('init_observation failed!')
            return False
        self.init_action_space()
        self.reward = 0.
        self.select_count = 0;
        self.select_index = 0
        self.select_total_reward = 0.
        self.state = np.array([self.reward,self.select_total_reward])
        self.result.clear()
        self.result_keywords.clear()
#        self.buf_obs[None][0] = self.observation.copy()
        return self.observation.copy()
    
    def seed(self,seed=None):
        self.np_random,seed = seeding.np_random(seed)
    
    def copy_obs_dict(self,obs):
        """
        Deep-copy an observation dict.
        """
        return {k: np.copy(v) for k, v in obs.items()}


    def dict_to_obs(self,obs_dict):
        """
        Convert an observation dict into a raw array if the
        original observation space was not a Dict space.
        """
        if set(obs_dict.keys()) == {None}:
            return obs_dict[None]
        return obs_dict
    
    def init_observation_highroi(self):
        logger.debug('observation-------------------------')
#        self.keywords_length = 10
        self.keywords_length = len(self.data)
        if self.keywords_length == 0:
            logger.info('self.keywords_length num is 0')
            return False
        logger.debug('self.keywords_length: '+ str(self.keywords_length))
        self.observation = np.empty(self.keywords_length * self.parameter_size,float)
        buy_num_max = self.data[0][16]
        buy_num_min = self.data[0][16]
        money_num_max = self.data[0][18]
        money_num_min = self.data[0][18]
        roi_max = self.data[0][19]
        roi_min = self.data[0][19]
        for i in range(self.keywords_length):
            self.observation[i*self.parameter_size] = self.data[i][16]
            if self.observation[i*self.parameter_size] > buy_num_max:
                buy_num_max = self.observation[i*self.parameter_size]
            if self.observation[i*self.parameter_size] < buy_num_min:
                buy_num_min = self.observation[i*self.parameter_size]
                
            self.observation[i*self.parameter_size + 1] = self.data[i][18]
            if self.observation[i*self.parameter_size + 1] > money_num_max:
                money_num_max = self.observation[i*self.parameter_size + 1]
            if self.observation[i*self.parameter_size + 1] < money_num_min:
                money_num_min = self.observation[i*self.parameter_size + 1]
                
            self.observation[i*self.parameter_size + 2] = self.data[i][19]
            if self.observation[i*self.parameter_size + 2] > roi_max:
                roi_max = self.observation[i*self.parameter_size + 2]
            if self.observation[i*self.parameter_size + 2] < roi_min:
                roi_min = self.observation[i*self.parameter_size + 2]
                

        #normal action
        normal_check_buy_num = 0.
        normal_check_money_num = 0.
        normal_check_roi_num = 0.
        logger.debug('buy_num_max:' + str(buy_num_max))
        logger.debug('buy_num_min:' + str(buy_num_min))
        logger.debug('money_num_max:' + str(money_num_max))
        logger.debug('money_num_min:' + str(money_num_min))
        logger.debug('roi_max:' + str(roi_max))
        logger.debug('roi_min:' + str(roi_min))

        for j in range(self.keywords_length):
            self.observation[j*self.parameter_size] = (self.observation[j*self.parameter_size] - buy_num_min) / (buy_num_max - buy_num_min)
            normal_check_buy_num+=self.observation[j*self.parameter_size]
            
            self.observation[j*self.parameter_size + 1] = (self.observation[j*self.parameter_size + 1] - money_num_min) / (money_num_max - money_num_min)
            normal_check_money_num+=self.observation[j*self.parameter_size + 1]
            
            self.observation[j*self.parameter_size + 2] = (self.observation[j*self.parameter_size + 2] - roi_min) / (roi_max - roi_min)
            normal_check_roi_num+=self.observation[j*self.parameter_size + 2]
                
        
        nomal_check_total = normal_check_buy_num + normal_check_money_num + normal_check_roi_num
        logger.debug('nomal_check_total:' + str(nomal_check_total))
        #add mean value
        normal_check_mean = nomal_check_total / self.observation.shape[0]
        logger.debug('normal_check_mean:' + str(normal_check_mean))
        normal_check_max = self.observation[0]
        normal_check_min = self.observation[0]
        for n in range(self.observation.shape[0]):
            self.observation[n] += normal_check_mean
#            logger.debug('self.observation[n]:' + str(self.observation[n]))
            if self.observation[n] > normal_check_max:
                normal_check_max = self.observation[n]
            if self.observation[n] <= normal_check_min:
                normal_check_min = self.observation[n]
        
        logger.debug('normal_check_max:' + str(normal_check_max))
        logger.debug('normal_check_min:' + str(normal_check_min))
        
        #check normal_action
        normal_check_last = 0.
        for x in range(self.observation.shape[0]):
#            logger.debug('self.observation[n]:' + str(self.observation[x]))
            self.observation[x] = (self.observation[x] - normal_check_min) / (normal_check_max - normal_check_min)
#            logger.debug('self.observation[n]_:' + str(self.observation[x]))
            normal_check_last += self.observation[x]
        logger.debug('normal_check_last:' + str(normal_check_last))
                    
        logger.debug('print total observation--------------------------')
        for a in range(self.keywords_length):
            print_data = []
            print_data.append(self.observation[a*self.parameter_size])
            print_data.append(self.observation[a*self.parameter_size + 1])
            print_data.append(self.observation[a*self.parameter_size + 2])
            logger.debug(str(print_data))
        logger.debug('end total observation--------------------------')
                    
#        logger.debug(str(self.observation))
        logger.debug('normal_check_buy_num:' + str(normal_check_buy_num))
        logger.debug('normal_check_money_num:' + str(normal_check_money_num))
        logger.debug('normal_check_roi_num:' + str(normal_check_roi_num))
        logger.debug('end_observation-------------------------')
        self.observation_space = Box(low = -self.observation,high = self.observation,dtype = np.float32)
#        self.keys, shapes, dtypes = self.obs_space_info(self.observation_space)
#        self.buf_obs = { k: np.zeros((1,) + tuple(shapes[k]), dtype=dtypes[k]) for k in self.keys }
#        self.buf_obs[None][0] = self.observation
        logger.debug('self.observation_space:')
        logger.debug(str(self.observation_space))
        return self.observation
    
    def init_observation_highflow(self):
        logger.debug('observation-------------------------')
#        self.keywords_length = 10
        self.keywords_length = len(self.data)
        if self.keywords_length == 0:
            logger.info('self.keywords_length num is 0')
            return False
        logger.debug('self.keywords_length: '+ str(self.keywords_length))
        self.observation = np.empty(self.keywords_length * self.parameter_size,float)
        popularity_max = self.data[0][1]
        popularity_min = self.data[0][1]
        click_num_max = self.data[0][12]
        click_num_min = self.data[0][12]
        click_rate_max = self.data[0][13]
        click_rate_min = self.data[0][13]
        for i in range(self.keywords_length):
            self.observation[i*self.parameter_size] = self.data[i][1]
            if self.observation[i*self.parameter_size] > popularity_max:
                popularity_max = self.observation[i*self.parameter_size]
            if self.observation[i*self.parameter_size] < popularity_min:
                popularity_min = self.observation[i*self.parameter_size]
                
            self.observation[i*self.parameter_size + 1] = self.data[i][12]
            if self.observation[i*self.parameter_size + 1] > click_num_max:
                click_num_max = self.observation[i*self.parameter_size + 1]
            if self.observation[i*self.parameter_size + 1] < click_num_min:
                click_num_min = self.observation[i*self.parameter_size + 1]
                
            self.observation[i*self.parameter_size + 2] = self.data[i][13]
            if self.observation[i*self.parameter_size + 2] > click_rate_max:
                click_rate_max = self.observation[i*self.parameter_size + 2]
            if self.observation[i*self.parameter_size + 2] < click_rate_min:
                click_rate_min = self.observation[i*self.parameter_size + 2]
                

        #normal action
        normal_check_popularity = 0.
        normal_check_click_num = 0.
        normal_check_click_rate = 0.
        logger.debug('popularity_max:' + str(popularity_max))
        logger.debug('popularity_min:' + str(popularity_min))
        logger.debug('click_num_max:' + str(click_num_max))
        logger.debug('click_num_min:' + str(click_num_min))
        logger.debug('click_rate_max:' + str(click_rate_max))
        logger.debug('click_rate_min:' + str(click_rate_min))

        for j in range(self.keywords_length):
            self.observation[j*self.parameter_size] = (self.observation[j*self.parameter_size] - popularity_min) / (popularity_max - popularity_min)
            normal_check_popularity+=self.observation[j*self.parameter_size]
            
            self.observation[j*self.parameter_size + 1] = (self.observation[j*self.parameter_size + 1] - click_num_min) / (click_num_max - click_num_min)
            normal_check_click_num+=self.observation[j*self.parameter_size + 1]
            
            self.observation[j*self.parameter_size + 2] = (self.observation[j*self.parameter_size + 2] - click_rate_min) / (click_rate_max - click_rate_min)
            normal_check_click_rate+=self.observation[j*self.parameter_size + 2]
                
        
        nomal_check_total = normal_check_popularity + normal_check_click_num + normal_check_click_rate
        logger.debug('nomal_check_total:' + str(nomal_check_total))
        #add mean value
        normal_check_mean = nomal_check_total / self.observation.shape[0]
        logger.debug('normal_check_mean:' + str(normal_check_mean))
        normal_check_max = self.observation[0]
        normal_check_min = self.observation[0]
        for n in range(self.observation.shape[0]):
            self.observation[n] += normal_check_mean
#            logger.debug('self.observation[n]:' + str(self.observation[n]))
            if self.observation[n] > normal_check_max:
                normal_check_max = self.observation[n]
            if self.observation[n] <= normal_check_min:
                normal_check_min = self.observation[n]
        
        logger.debug('normal_check_max:' + str(normal_check_max))
        logger.debug('normal_check_min:' + str(normal_check_min))
        
        #check normal_action
        normal_check_last = 0.
        for x in range(self.observation.shape[0]):
#            logger.debug('self.observation[n]:' + str(self.observation[x]))
            self.observation[x] = (self.observation[x] - normal_check_min) / (normal_check_max - normal_check_min)
#            logger.debug('self.observation[n]_:' + str(self.observation[x]))
            normal_check_last += self.observation[x]
        logger.debug('normal_check_last:' + str(normal_check_last))
                    
        logger.debug('print total observation--------------------------')
        for a in range(self.keywords_length):
            print_data = []
            print_data.append(self.observation[a*self.parameter_size])
            print_data.append(self.observation[a*self.parameter_size + 1])
            print_data.append(self.observation[a*self.parameter_size + 2])
            logger.debug(str(print_data))
        logger.debug('end total observation--------------------------')
                    
#        logger.debug(str(self.observation))
        logger.debug('normal_check_popularity:' + str(normal_check_popularity))
        logger.debug('normal_check_click_num:' + str(normal_check_click_num))
        logger.debug('normal_check_click_rate:' + str(normal_check_click_rate))
        logger.debug('end_observation-------------------------')
        self.observation_space = Box(low = -self.observation,high = self.observation,dtype = np.float32)
#        self.keys, shapes, dtypes = self.obs_space_info(self.observation_space)
#        self.buf_obs = { k: np.zeros((1,) + tuple(shapes[k]), dtype=dtypes[k]) for k in self.keys }
#        self.buf_obs[None][0] = self.observation
        logger.debug('self.observation_space:')
        logger.debug(str(self.observation_space))
        return self.observation
    
    def init_observation_highroi_hightconveration(self):
        logger.debug('observation-------------------------')
#        self.keywords_length = 10
        self.keywords_length = len(self.data)
        if self.keywords_length == 0:
            logger.info('self.keywords_length num is 0')
            return False
        logger.debug('self.keywords_length: '+ str(self.keywords_length))
        self.observation = np.empty(self.keywords_length * self.parameter_size,float)
        popularity_max = self.data[0][1]
        popularity_min = self.data[0][1]
        conversion_max = self.data[0][2]
        conversion_min = self.data[0][2]
        transform_1_max = self.data[0][9]
        transform_1_min = self.data[0][9]
        transform_2_max = self.data[0][10]
        transform_2_min = self.data[0][10]
        transform_3_max = self.data[0][11]
        transform_3_min = self.data[0][11]
        buy_num_max = self.data[0][16]
        buy_num_min = self.data[0][16]
        money_num_max = self.data[0][18]
        money_num_min = self.data[0][18]
        roi_max = self.data[0][19]
        roi_min = self.data[0][19]
        for i in range(self.keywords_length):
            self.observation[i*self.parameter_size] = self.data[i][1]
            if self.observation[i*self.parameter_size] > popularity_max:
                popularity_max = self.observation[i*self.parameter_size]
            if self.observation[i*self.parameter_size] < popularity_min:
                popularity_min = self.observation[i*self.parameter_size]
                
            self.observation[i*self.parameter_size + 1] = self.data[i][2]
            if self.observation[i*self.parameter_size + 1] > conversion_max:
                conversion_max = self.observation[i*self.parameter_size + 1]
            if self.observation[i*self.parameter_size + 1] < conversion_min:
                conversion_min = self.observation[i*self.parameter_size + 1]
                
            self.observation[i*self.parameter_size + 2] = self.data[i][9]
            if self.observation[i*self.parameter_size + 2] > transform_1_max:
                transform_1_max = self.observation[i*self.parameter_size + 2]
            if self.observation[i*self.parameter_size + 2] < transform_1_min:
                transform_1_min = self.observation[i*self.parameter_size + 2]
                
            self.observation[i*self.parameter_size + 3] = self.data[i][10]
            if self.observation[i*self.parameter_size + 3] > transform_2_max:
                transform_2_max = self.observation[i*self.parameter_size + 3]
            if self.observation[i*self.parameter_size + 3] < transform_2_min:
                transform_2_min = self.observation[i*self.parameter_size + 3]
                
            self.observation[i*self.parameter_size + 4] = self.data[i][11]
            if self.observation[i*self.parameter_size + 4] > transform_3_max:
                transform_3_max = self.observation[i*self.parameter_size + 4]
            if self.observation[i*self.parameter_size + 4] < transform_3_min:
                transform_3_min = self.observation[i*self.parameter_size + 4]
                
            self.observation[i*self.parameter_size + 5] = self.data[i][16]
            if self.observation[i*self.parameter_size] > buy_num_max:
                buy_num_max = self.observation[i*self.parameter_size + 5]
            if self.observation[i*self.parameter_size] < buy_num_min:
                buy_num_min = self.observation[i*self.parameter_size + 5]
                
            self.observation[i*self.parameter_size + 6] = self.data[i][18]
            if self.observation[i*self.parameter_size + 6] > money_num_max:
                money_num_max = self.observation[i*self.parameter_size + 6]
            if self.observation[i*self.parameter_size + 6] < money_num_min:
                money_num_min = self.observation[i*self.parameter_size + 6]
                
            self.observation[i*self.parameter_size + 7] = self.data[i][19]
            if self.observation[i*self.parameter_size + 7] > roi_max:
                roi_max = self.observation[i*self.parameter_size + 7]
            if self.observation[i*self.parameter_size + 7] < roi_min:
                roi_min = self.observation[i*self.parameter_size + 7]

        #normal action
        normal_check_popularity = 0.
        normal_check_conversion = 0.
        normal_check_transform_1 = 0.
        normal_check_transform_2 = 0.
        normal_check_transform_3 = 0.
        
        normal_check_buy_num = 0.
        normal_check_money_num = 0.
        normal_check_roi_num = 0.
        
        logger.debug('buy_num_max:' + str(buy_num_max))
        logger.debug('buy_num_min:' + str(buy_num_min))
        logger.debug('money_num_max:' + str(money_num_max))
        logger.debug('money_num_min:' + str(money_num_min))
        logger.debug('roi_max:' + str(roi_max))
        logger.debug('roi_min:' + str(roi_min))
        logger.debug('popularity_max:' + str(popularity_max))
        logger.debug('popularity_min:' + str(popularity_min))
        logger.debug('conversion_max:' + str(conversion_max))
        logger.debug('conversion_min:' + str(conversion_min))
        logger.debug('transform_1_max:' + str(transform_1_max))
        logger.debug('transform_1_min:' + str(transform_1_min))
        logger.debug('transform_2_max:' + str(transform_2_max))
        logger.debug('transform_2_min:' + str(transform_2_min))
        logger.debug('transform_3_max:' + str(transform_3_max))
        logger.debug('transform_3_min:' + str(transform_3_min))
        for j in range(self.keywords_length):
            self.observation[j*self.parameter_size] = (self.observation[j*self.parameter_size] - popularity_min) / (popularity_max - popularity_min)
            normal_check_popularity+=self.observation[j*self.parameter_size]
            
            self.observation[j*self.parameter_size + 1] = (self.observation[j*self.parameter_size + 1] - conversion_min) / (conversion_max - conversion_min)
            normal_check_conversion+=self.observation[j*self.parameter_size + 1]
            
            self.observation[j*self.parameter_size + 2] = (self.observation[j*self.parameter_size + 2] - transform_1_min) / (transform_1_max - transform_1_min)
            normal_check_transform_1+=self.observation[j*self.parameter_size + 2]
            
            self.observation[j*self.parameter_size + 3] = (self.observation[j*self.parameter_size + 3] - transform_2_min) / (transform_2_max - transform_2_min)
            normal_check_transform_2+=self.observation[j*self.parameter_size + 3]
            
            self.observation[j*self.parameter_size + 4] = (self.observation[j*self.parameter_size + 4] - transform_3_min) / (transform_3_max - transform_3_min)
            normal_check_transform_3+=self.observation[j*self.parameter_size + 4]
            
            self.observation[j*self.parameter_size + 5] = (self.observation[j*self.parameter_size + 5] - buy_num_min) / (buy_num_max - buy_num_min)
            normal_check_buy_num+=self.observation[j*self.parameter_size]
            
            self.observation[j*self.parameter_size + 6] = (self.observation[j*self.parameter_size + 6] - money_num_min) / (money_num_max - money_num_min)
            normal_check_money_num+=self.observation[j*self.parameter_size + 1]
            
            self.observation[j*self.parameter_size + 7] = (self.observation[j*self.parameter_size + 7] - roi_min) / (roi_max - roi_min)
            normal_check_roi_num+=self.observation[j*self.parameter_size + 2]
            
        
        nomal_check_total = normal_check_popularity + normal_check_conversion + normal_check_transform_1 + normal_check_transform_2 + normal_check_transform_3 + normal_check_buy_num + normal_check_money_num + normal_check_roi_num
        logger.debug('nomal_check_total:' + str(nomal_check_total))
        #add mean value
        normal_check_mean = nomal_check_total / self.observation.shape[0]
        logger.debug('normal_check_mean:' + str(normal_check_mean))
        normal_check_max = self.observation[0]
        normal_check_min = self.observation[0]
        for n in range(self.observation.shape[0]):
            self.observation[n] += normal_check_mean
#            logger.debug('self.observation[n]:' + str(self.observation[n]))
            if self.observation[n] > normal_check_max:
                normal_check_max = self.observation[n]
            if self.observation[n] <= normal_check_min:
                normal_check_min = self.observation[n]
        
        logger.debug('normal_check_max:' + str(normal_check_max))
        logger.debug('normal_check_min:' + str(normal_check_min))
        
        #check normal_action
        normal_check_last = 0.
        for x in range(self.observation.shape[0]):
#            logger.debug('self.observation[n]:' + str(self.observation[x]))
            self.observation[x] = (self.observation[x] - normal_check_min) / (normal_check_max - normal_check_min)
#            logger.debug('self.observation[n]_:' + str(self.observation[x]))
            normal_check_last += self.observation[x]
        logger.debug('normal_check_last:' + str(normal_check_last))
                    
        logger.debug('print total observation--------------------------')
        for a in range(self.keywords_length):
            print_data = []
            print_data.append(self.observation[a*self.parameter_size])
            print_data.append(self.observation[a*self.parameter_size + 1])
            print_data.append(self.observation[a*self.parameter_size + 2])
            print_data.append(self.observation[a*self.parameter_size + 3])
            print_data.append(self.observation[a*self.parameter_size + 4])
            print_data.append(self.observation[a*self.parameter_size + 5])
            print_data.append(self.observation[a*self.parameter_size + 6])
            print_data.append(self.observation[a*self.parameter_size + 7])
            logger.debug(str(print_data))
        logger.debug('end total observation--------------------------')
                    
#        logger.debug(str(self.observation))
        logger.debug('normal_check_popularity:' + str(normal_check_popularity))
        logger.debug('normal_check_conversion:' + str(normal_check_conversion))
        logger.debug('normal_check_transform_1:' + str(normal_check_transform_1))
        logger.debug('normal_check_transform_2:' + str(normal_check_transform_2))
        logger.debug('normal_check_transform_3:' + str(normal_check_transform_3))
        logger.debug('normal_check_buy_num:' + str(normal_check_buy_num))
        logger.debug('normal_check_money_num:' + str(normal_check_money_num))
        logger.debug('normal_check_roi_num:' + str(normal_check_roi_num))
        logger.debug('end_observation-------------------------')
        self.observation_space = Box(low = -self.observation,high = self.observation,dtype = np.float32)
#        self.keys, shapes, dtypes = self.obs_space_info(self.observation_space)
#        self.buf_obs = { k: np.zeros((1,) + tuple(shapes[k]), dtype=dtypes[k]) for k in self.keys }
#        self.buf_obs[None][0] = self.observation
        logger.debug('self.observation_space:')
        logger.debug(str(self.observation_space))
        return self.observation
    
    def init_observation(self):
        logger.debug('observation-------------------------')
#        self.keywords_length = 10
        self.keywords_length = len(self.data)
        if self.keywords_length == 0:
            logger.info('self.keywords_length num is 0')
            return False
        logger.debug('self.keywords_length: '+ str(self.keywords_length))
        self.observation = np.empty(self.keywords_length * self.parameter_size,float)
        popularity_max = self.data[0][1]
        popularity_min = self.data[0][1]
        conversion_max = self.data[0][2]
        conversion_min = self.data[0][2]
        transform_1_max = self.data[0][9]
        transform_1_min = self.data[0][9]
        transform_2_max = self.data[0][10]
        transform_2_min = self.data[0][10]
        transform_3_max = self.data[0][11]
        transform_3_min = self.data[0][11]
        for i in range(self.keywords_length):
            self.observation[i*self.parameter_size] = self.data[i][1]
            if self.observation[i*self.parameter_size] > popularity_max:
                popularity_max = self.observation[i*self.parameter_size]
            if self.observation[i*self.parameter_size] < popularity_min:
                popularity_min = self.observation[i*self.parameter_size]
                
            self.observation[i*self.parameter_size + 1] = self.data[i][2]
            if self.observation[i*self.parameter_size + 1] > conversion_max:
                conversion_max = self.observation[i*self.parameter_size + 1]
            if self.observation[i*self.parameter_size + 1] < conversion_min:
                conversion_min = self.observation[i*self.parameter_size + 1]
                
            self.observation[i*self.parameter_size + 2] = self.data[i][9]
            if self.observation[i*self.parameter_size + 2] > transform_1_max:
                transform_1_max = self.observation[i*self.parameter_size + 2]
            if self.observation[i*self.parameter_size + 2] < transform_1_min:
                transform_1_min = self.observation[i*self.parameter_size + 2]
                
            self.observation[i*self.parameter_size + 3] = self.data[i][10]
            if self.observation[i*self.parameter_size + 3] > transform_2_max:
                transform_2_max = self.observation[i*self.parameter_size + 3]
            if self.observation[i*self.parameter_size + 3] < transform_2_min:
                transform_2_min = self.observation[i*self.parameter_size + 3]
                
            self.observation[i*self.parameter_size + 4] = self.data[i][11]
            if self.observation[i*self.parameter_size + 4] > transform_3_max:
                transform_3_max = self.observation[i*self.parameter_size + 4]
            if self.observation[i*self.parameter_size + 4] < transform_3_min:
                transform_3_min = self.observation[i*self.parameter_size + 4]

        #normal action
        normal_check_popularity = 0.
        normal_check_conversion = 0.
        normal_check_transform_1 = 0.
        normal_check_transform_2 = 0.
        normal_check_transform_3 = 0.
        logger.debug('popularity_max:' + str(popularity_max))
        logger.debug('popularity_min:' + str(popularity_min))
        logger.debug('conversion_max:' + str(conversion_max))
        logger.debug('conversion_min:' + str(conversion_min))
        logger.debug('transform_1_max:' + str(transform_1_max))
        logger.debug('transform_1_min:' + str(transform_1_min))
        logger.debug('transform_2_max:' + str(transform_2_max))
        logger.debug('transform_2_min:' + str(transform_2_min))
        logger.debug('transform_3_max:' + str(transform_3_max))
        logger.debug('transform_3_min:' + str(transform_3_min))
        for j in range(self.keywords_length):
            self.observation[j*self.parameter_size] = (self.observation[j*self.parameter_size] - popularity_min) / (popularity_max - popularity_min)
            normal_check_popularity+=self.observation[j*self.parameter_size]
            
            self.observation[j*self.parameter_size + 1] = (self.observation[j*self.parameter_size + 1] - conversion_min) / (conversion_max - conversion_min)
            normal_check_conversion+=self.observation[j*self.parameter_size + 1]
            
            self.observation[j*self.parameter_size + 2] = (self.observation[j*self.parameter_size + 2] - transform_1_min) / (transform_1_max - transform_1_min)
            normal_check_transform_1+=self.observation[j*self.parameter_size + 2]
            
            self.observation[j*self.parameter_size + 3] = (self.observation[j*self.parameter_size + 3] - transform_2_min) / (transform_2_max - transform_2_min)
            normal_check_transform_2+=self.observation[j*self.parameter_size + 3]
            
            self.observation[j*self.parameter_size + 4] = (self.observation[j*self.parameter_size + 4] - transform_3_min) / (transform_3_max - transform_3_min)
            normal_check_transform_3+=self.observation[j*self.parameter_size + 4]
            
        
        nomal_check_total = normal_check_popularity + normal_check_conversion + normal_check_transform_1 + normal_check_transform_2 + normal_check_transform_3
        logger.debug('nomal_check_total:' + str(nomal_check_total))
        #add mean value
        normal_check_mean = nomal_check_total / self.observation.shape[0]
        logger.debug('normal_check_mean:' + str(normal_check_mean))
        normal_check_max = self.observation[0]
        normal_check_min = self.observation[0]
        for n in range(self.observation.shape[0]):
            self.observation[n] += normal_check_mean
#            logger.debug('self.observation[n]:' + str(self.observation[n]))
            if self.observation[n] > normal_check_max:
                normal_check_max = self.observation[n]
            if self.observation[n] <= normal_check_min:
                normal_check_min = self.observation[n]
        
        logger.debug('normal_check_max:' + str(normal_check_max))
        logger.debug('normal_check_min:' + str(normal_check_min))
        
        #check normal_action
        normal_check_last = 0.
        for x in range(self.observation.shape[0]):
#            logger.debug('self.observation[n]:' + str(self.observation[x]))
            self.observation[x] = (self.observation[x] - normal_check_min) / (normal_check_max - normal_check_min)
#            logger.debug('self.observation[n]_:' + str(self.observation[x]))
            normal_check_last += self.observation[x]
        logger.debug('normal_check_last:' + str(normal_check_last))
                    
        logger.debug('print total observation--------------------------')
        for a in range(self.keywords_length):
            print_data = []
            print_data.append(self.observation[a*self.parameter_size])
            print_data.append(self.observation[a*self.parameter_size + 1])
            print_data.append(self.observation[a*self.parameter_size + 2])
            print_data.append(self.observation[a*self.parameter_size + 3])
            print_data.append(self.observation[a*self.parameter_size + 4])
            logger.debug(str(print_data))
        logger.debug('end total observation--------------------------')
                    
#        logger.debug(str(self.observation))
        logger.debug('normal_check_popularity:' + str(normal_check_popularity))
        logger.debug('normal_check_conversion:' + str(normal_check_conversion))
        logger.debug('normal_check_transform_1:' + str(normal_check_transform_1))
        logger.debug('normal_check_transform_2:' + str(normal_check_transform_2))
        logger.debug('normal_check_transform_3:' + str(normal_check_transform_3))
        logger.debug('end_observation-------------------------')
        self.observation_space = Box(low = -self.observation,high = self.observation,dtype = np.float32)
#        self.keys, shapes, dtypes = self.obs_space_info(self.observation_space)
#        self.buf_obs = { k: np.zeros((1,) + tuple(shapes[k]), dtype=dtypes[k]) for k in self.keys }
#        self.buf_obs[None][0] = self.observation
        logger.debug('self.observation_space:')
        logger.debug(str(self.observation_space))
        return self.observation
    
    def get_observation(self):
        return self.observation.copy()
    
    def init_action_space(self):
        self.max_action = 1.
        self.action_space = Box(low = self.max_action,high = self.max_action,shape=(1,),dtype = np.float32)
        logger.debug('self.action_space:')
        logger.debug(str(self.action_space))
        logger.debug('self.action_space.shape[-1]:'+str(self.action_space.shape[-1]))
    
    def get_action_space(self):
        return self.action_space
    
    def get_result_keywords(self):
        logger.debug('--------------get_result_keywords----------------')
        self.result_keywords.clear()
        count = len(self.result)
        self.result.sort()
        logger.debug('self.result: ' + str(self.result))
        logger.debug('result count: ' + str(count))
        for i in range(count):
            key_id = int(self.result[i])
            key_words = self.data[key_id][0]
            self.result_keywords.append(key_words)
            logger.debug('key_id: ' + str(key_id))
            logger.debug('keywords: ' + str(key_words))
        
        logger.debug('result keywords: ' + str(self.result_keywords))
        logger.debug('--------------end get_result_keywords----------------')
        return self.result_keywords
    
    def step(self,u):
        logger.debug('wordAgent select: ' + str(u))
        a_value =  float(u[0])
        pervalue = (a_value + 1.) / 2.
        logger.debug('wordAgent pervalue: ' + str(pervalue))
        index = int(pervalue * float(self.keywords_length - 1))
        logger.debug('--------------step----------------')
        logger.debug('wordAgent select index: ' + str(index))
        logger.debug('step self.observation index: ' + str(index))
        self.select_index = index;
        popularity = self.observation[index*self.parameter_size]
        logger.debug('step popularity: ' + str(popularity))
        keywords_id = index
        
        self.reward = agent.reward.get_reward_value(self.observation,self.parameter_size,index,self.reward_type)
        
        self.observation[index*self.parameter_size] = -1.
#        self.reward = popularity*conversion + popularity*transform_1*2 + popularity*transform_2 + popularity*transform_3
        if popularity == -1.:
            self.reward = -1
        self.reward *= 10
        logger.debug('wordAgent self.reward: ' + str(self.reward))
        self.result.append(int(keywords_id))
        self.select_total_reward += self.reward
        
        logger.debug('step self.observation result: ' + str(self.result))
        logger.debug('----------------step end---------------- ')
        self.state[0] = self.reward
        self.state[1] = self.select_total_reward
        self.select_count += 1
        done = False
        if self.select_count >= self.all_words:
            done = True
#        self.buf_obs[None][0] = self.observation.copy()
        return self.observation.copy(),self.reward,done,{}
    
    def render(self,mode = 'human'):
         
#        logger.info('------------------------------------------')
        renderdic = {}
        renderdic['current_reward'] = self.reward
        renderdic['select_total_reward'] = self.select_total_reward
        renderdic['result_idlist'] = self.result
        renderdic['result_keywords'] = self.get_result_keywords()
        renderdic['select_index'] =  self.select_index
#        logger.info('')
#        logger.info(str(renderdic))
#        logger.info('')
#        logger.info('------------------------------------------')
    
    
    def get_state(self):
        return self.state
    
    def obs_space_info(self,obs_space):
        """
        Get dict-structured information about a gym.Space.

        Returns:
            A tuple (keys, shapes, dtypes):
                keys: a list of dict keys.
                shapes: a dict mapping keys to shapes.
                dtypes: a dict mapping keys to dtypes.
                """
        if isinstance(obs_space, Dict):
            assert isinstance(obs_space.spaces, OrderedDict)
            subspaces = obs_space.spaces
        else:
            subspaces = {None: obs_space}
        keys = []
        shapes = {}
        dtypes = {}
        for key, box in subspaces.items():
            keys.append(key)
            shapes[key] = box.shape
            dtypes[key] = box.dtype
        return keys, shapes, dtypes

#logger.debug('test openpyxl sucessful!')
#wordimp = WordAgent('../assert/collagen.xlsx','xlsx',COLLAGE_HIGHCONVERSION)

    
    
    